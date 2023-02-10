import openpyxl
from loguru import logger
import json
import os


class Trans_utils(object):

    def __init__(self, filename):
        self.excel_file = filename
        print("要转化的excel文件是：",self.excel_file)
        # self.book = xlrd2.open_workbook(self.excel_file)
        self.workbook=openpyxl.load_workbook(self.excel_file) # 打开文件，获取workbook对象
        self.sheet = self.workbook[self.workbook.sheetnames[0]]  # 获取sheet表单对象
        self.row = self.sheet.max_row  # 获取总行数
        print("总行数为", self.row)

     # 写入json文件
    def write_json(self,case, filename):
        with open(filename, "w", encoding='utf8') as f:
            logger.info("正在写入json文件......")
            # 加上4个空格；False 不转义
            json.dump(case, f, indent=4, ensure_ascii=False)

    # 读取excel文件并生成json文件
    def get_excel_data(self):
        logger.info("------------开始读取excel用例数据......------------")
         # excel对应列数
        cell_header = {
        "case_name":1,
        "is_run": 2,
        "headers": 3,
        "url": 4,
        "method": 5,
        "params_type": 6,
        "data": 7,
        "sql":8,
        "expect": 9
    }
         # 新建一个空列表（存储每行的数据）
        case = list()
        # 遍历每行数据
        for i in range(2, self.row + 1):
            # 新建空字典（存储每行数据）
            data = dict()
            # 判断是否执行
            if self.sheet.cell(i, cell_header.get("is_run")).value == "是":
                try:
                    # 读取数据追加到字典
                    data['case_name'] = self.sheet.cell(i, cell_header.get("case_name")).value
                    data['url'] = self.sheet.cell(i, cell_header.get("url")).value
                    data['method'] = self.sheet.cell(i, cell_header.get("method")).value
                    data['params_type'] = self.sheet.cell(i, cell_header.get("params_type")).value
                    data['headers'] = json.loads(self.sheet.cell(i, cell_header.get("headers")).value)  #数据json字符串转化为字典
                    data['data'] = eval(self.sheet.cell(i, cell_header.get("data")).value)
                    # 此处预期结果用str(), json.loads()或者eval() 处理，将json格式转为str，以实际情况选择对应函数
                    data['sql'] = str(self.sheet.cell(i, cell_header.get("sql")).value)
                    data['expect'] = str(self.sheet.cell(i, cell_header.get("expect")).value)
                    # 将字典追加到列表
                    case.append(data)

                except Exception as e:
                    # self.write_excel([i, cell_config.get("desc")], e)
                    logger.info(f"第{i}行读取失败，请检查数据格式是否正确！{e}！")
            # print(data)
        logger.info("------------excel数据读取结束！------------")
        # 将列表数据写入json
        filename0=os.path.basename(self.excel_file)
        filename=filename0.split(".")
        self.write_json(case, "../data/case_json/"+filename[0]+".json")
        logger.info("用例已生成json文件！")



if __name__=='__main__':

    data1=Trans_utils("D:\MySoftware\AutoTest\\apiAuto1\data\case_data.xlsx").get_excel_data()
    print(data1)